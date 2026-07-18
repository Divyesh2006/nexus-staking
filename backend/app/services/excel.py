from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import StakingRecord

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1D4ED8")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
CENTER = Alignment(horizontal="center", vertical="center")

COLUMNS = [
    "USERNAME",
    "STAKING ID",
    "MODE",
    "VOLUME",
    "ROI",
    "ROI REWARDS",
    "STAKING DATE",
    "MATURITY DATE",
    "STAKING YEARS",
    "MATURITY",
    "MATURITY REWARD",
]


def build_excel_workbook(records: list[StakingRecord], output_dir: Path) -> str:
    if not records:
        raise ValueError("records cannot be empty")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Nexus Stakings"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:K{len(records) + 1}"

    for column_index, column_name in enumerate(COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row_index, record in enumerate(records, start=2):
        values = [
            record.username,
            record.staking_id,
            record.mode,
            record.volume,
            record.roi if record.roi is not None else "",
            record.staking_date,
            record.maturity_date,
            record.staking_years,
            record.maturity,
            record.maturity_reward,
        ]
        # Insert roi_rewards after ROI
        values.insert(5, record.roi_rewards if record.roi_rewards is not None else "")
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if column_index in {4, 9, 10, 5}:
                cell.number_format = '#,##0'
            if column_index == 5:
                # ROI may be a string (invalid message) or a float
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '@'
            if column_index in {6, 7} and hasattr(value, 'strftime'):
                cell.number_format = 'mm/dd/yyyy'

    for column_index, column_name in enumerate(COLUMNS, start=1):
        max_length = len(column_name)
        for row in worksheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
            cell = row[0]
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 4, 24)

    output_dir.mkdir(parents=True, exist_ok=True)
    filename_root = records[0].username.strip() or "Nexus"
    safe_root = "".join(character for character in filename_root if character.isalnum() or character in {"_", "-"}) or "Nexus"
    filename = f"{safe_root}_Stakings.xlsx"
    filepath = output_dir / filename
    workbook.save(filepath)
    return str(filepath)
